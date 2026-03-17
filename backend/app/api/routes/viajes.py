from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user, is_cointra_admin
from app.db.session import get_db
from app.models.conciliacion import Conciliacion
from app.models.conciliacion_item import ConciliacionItem
from app.models.enums import UserRole
from app.models.enums import ItemTipo
from app.models.operacion import Operacion
from app.models.usuario import Usuario
from app.models.viaje import Viaje
from app.schemas.viaje import CargaMasivaResultado, ViajeCreate, ViajeOut, ViajeUpdate
from app.services.pricing import calculate_tarifa_cliente

router = APIRouter(prefix="/viajes", tags=["viajes"])


def _estado_conciliacion_valor(viaje: Viaje) -> str | None:
    if viaje.estado_conciliacion:
        return viaje.estado_conciliacion
    if not viaje.conciliacion:
        return None
    estado = viaje.conciliacion.estado
    return getattr(estado, "value", estado)


def _expected_conciliado(viaje: Viaje) -> bool:
    estado = _estado_conciliacion_valor(viaje)
    if estado is None:
        return False
    return estado in {"APROBADA", "CERRADA"}


def _validate_user_access_operacion(user: Usuario, operacion: Operacion) -> None:
    if user.rol == UserRole.CLIENTE and user.cliente_id != operacion.cliente_id:
        raise HTTPException(status_code=403, detail="Operacion no disponible para este cliente")
    if user.rol == UserRole.TERCERO and user.tercero_id != operacion.tercero_id:
        raise HTTPException(status_code=403, detail="Operacion no disponible para este tercero")


def _ensure_cointra_admin(user: Usuario) -> None:
    if not is_cointra_admin(user):
        raise HTTPException(status_code=403, detail="Solo COINTRA_ADMIN puede editar o inactivar viajes")


@router.post("", response_model=ViajeOut)
def create_viaje(
    payload: ViajeCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    # Crear viajes: COINTRA_ADMIN, COINTRA_USER, TERCERO
    if user.rol == UserRole.TERCERO:
        allowed = True
    elif user.rol == UserRole.COINTRA:
        # Cualquier sub_rol de Cointra puede crear viajes
        allowed = True
    else:
        allowed = False

    if not allowed:
        raise HTTPException(status_code=403, detail="No tiene permisos para crear viajes")

    operacion = db.get(Operacion, payload.operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    viaje = Viaje(
        operacion_id=payload.operacion_id,
        tercero_id=operacion.tercero_id,
        titulo=payload.titulo,
        fecha_servicio=payload.fecha_servicio,
        origen=payload.origen,
        destino=payload.destino,
        placa=payload.placa,
        conductor=payload.conductor,
        tarifa_tercero=payload.tarifa_tercero,
        tarifa_cliente=payload.tarifa_cliente,
        manifiesto_numero=payload.manifiesto_numero,
        descripcion=payload.descripcion,
        created_by=user.id,
        cargado_por=user.rol.value,
        estado_conciliacion=None,
        activo=True,
    )

    if viaje.tarifa_tercero and not viaje.tarifa_cliente:
        viaje.tarifa_cliente, viaje.rentabilidad = calculate_tarifa_cliente(float(viaje.tarifa_tercero), operacion)

    db.add(viaje)
    db.commit()
    db.refresh(viaje)
    return viaje


@router.get("", response_model=list[dict])
def list_viajes(
    operacion_id: int | None = None,
    only_pending: bool = False,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    query = (
        db.query(Viaje)
        .join(Operacion, Operacion.id == Viaje.operacion_id)
        .options(selectinload(Viaje.conciliacion))
    )
    if not is_cointra_admin(user):
        query = query.filter(Viaje.activo.is_(True))
    if user.rol == UserRole.CLIENTE and user.cliente_id:
        query = query.filter(Operacion.cliente_id == user.cliente_id)
    if user.rol == UserRole.TERCERO and user.tercero_id:
        query = query.filter(Operacion.tercero_id == user.tercero_id)
    if operacion_id:
        query = query.filter(Viaje.operacion_id == operacion_id)
    if only_pending:
        query = query.filter(Viaje.conciliado.is_(False))

    viajes = query.order_by(Viaje.id.desc()).all()

    changed = False

    # Repara enlaces historicos: item VIAJE existente sin viaje.conciliacion_id sincronizado.
    viaje_ids = [v.id for v in viajes]
    item_conciliacion_by_viaje_id: dict[int, int] = {}
    if viaje_ids:
        item_rows = (
            db.query(ConciliacionItem.viaje_id, ConciliacionItem.conciliacion_id)
            .filter(
                ConciliacionItem.tipo == ItemTipo.VIAJE,
                ConciliacionItem.viaje_id.in_(viaje_ids),
                ConciliacionItem.viaje_id.is_not(None),
            )
            .order_by(ConciliacionItem.id.desc())
            .all()
        )
        for viaje_id, conciliacion_id in item_rows:
            if viaje_id not in item_conciliacion_by_viaje_id:
                item_conciliacion_by_viaje_id[viaje_id] = conciliacion_id

    conciliacion_ids_from_items = set(item_conciliacion_by_viaje_id.values())
    conciliaciones_map: dict[int, str] = {}
    if conciliacion_ids_from_items:
        conciliaciones = (
            db.query(Conciliacion.id, Conciliacion.estado)
            .filter(Conciliacion.id.in_(conciliacion_ids_from_items))
            .all()
        )
        conciliaciones_map = {cid: getattr(estado, "value", estado) for cid, estado in conciliaciones}

    for viaje in viajes:
        recovered_conciliacion_id = item_conciliacion_by_viaje_id.get(viaje.id)
        effective_conciliacion_id = viaje.conciliacion_id or recovered_conciliacion_id
        if viaje.conciliacion_id is None and recovered_conciliacion_id is not None:
            viaje.conciliacion_id = recovered_conciliacion_id
            changed = True

        estado_actual = _estado_conciliacion_valor(viaje)
        if estado_actual is None and effective_conciliacion_id is not None:
            estado_actual = conciliaciones_map.get(effective_conciliacion_id)

        esperado = _expected_conciliado(viaje)
        if estado_actual is not None:
            esperado = estado_actual in {"APROBADA", "CERRADA"}

        if viaje.conciliado != esperado:
            viaje.conciliado = esperado
            changed = True
        if viaje.estado_conciliacion != estado_actual:
            viaje.estado_conciliacion = estado_actual
            changed = True

    if changed:
        db.commit()

    payload: list[dict] = []
    for viaje in viajes:
        effective_estado = viaje.estado_conciliacion
        if user.rol == UserRole.CLIENTE and effective_estado not in {"EN_REVISION", "APROBADA", "CERRADA"}:
            continue

        out = ViajeOut.model_validate(viaje).model_dump()
        if viaje.conciliacion_id is None and viaje.id in item_conciliacion_by_viaje_id:
            out["conciliacion_id"] = item_conciliacion_by_viaje_id[viaje.id]
        out["estado_conciliacion"] = effective_estado
        payload.append(out)

    return payload


@router.patch("/{viaje_id}", response_model=ViajeOut)
def update_viaje(
    viaje_id: int,
    payload: ViajeUpdate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    viaje = db.get(Viaje, viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No se enviaron cambios")

    if "tarifa_tercero" in data and data.get("tarifa_tercero") is not None:
        operacion = db.get(Operacion, viaje.operacion_id)
        tarifa_tercero = float(data["tarifa_tercero"])
        if "tarifa_cliente" not in data or data.get("tarifa_cliente") in (None, 0):
            tarifa_cliente, rentabilidad = calculate_tarifa_cliente(tarifa_tercero, operacion)
            data["tarifa_cliente"] = tarifa_cliente
            data["rentabilidad"] = rentabilidad

    for field, value in data.items():
        setattr(viaje, field, value)

    db.commit()
    db.refresh(viaje)
    return viaje


@router.delete("/{viaje_id}")
def deactivate_viaje(
    viaje_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    viaje = db.get(Viaje, viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    viaje.activo = False
    db.commit()
    return {"ok": True}


@router.post("/{viaje_id}/reactivar")
def reactivate_viaje(
    viaje_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    viaje = db.get(Viaje, viaje_id)
    if not viaje:
        raise HTTPException(status_code=404, detail="Viaje no encontrado")

    viaje.activo = True
    db.commit()
    return {"ok": True}


@router.post("/carga-masiva", response_model=CargaMasivaResultado)
async def bulk_upload_viajes(
    operacion_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    # Carga masiva de viajes: mismos permisos que crear viaje
    if user.rol == UserRole.TERCERO:
        allowed = True
    elif user.rol == UserRole.COINTRA:
        allowed = True
    else:
        allowed = False

    if not allowed:
        raise HTTPException(status_code=403, detail="No tiene permisos para cargar viajes")

    operacion = db.get(Operacion, operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Archivo Excel vacio")

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    expected = {
        "titulo",
        "fecha_servicio",
        "origen",
        "destino",
        "placa",
        "tarifa_tercero",
    }
    missing = [h for h in expected if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Columnas faltantes: {', '.join(missing)}")

    idx = {name: headers.index(name) for name in headers if name}
    errores: list[str] = []
    cargados = 0

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            titulo = (row[idx["titulo"]] or "").strip() if row[idx["titulo"]] else ""
            fecha_val = row[idx["fecha_servicio"]]
            origen = (row[idx["origen"]] or "").strip() if row[idx["origen"]] else ""
            destino = (row[idx["destino"]] or "").strip() if row[idx["destino"]] else ""
            placa = (row[idx["placa"]] or "").strip() if row[idx["placa"]] else ""
            conductor = ""
            if "conductor" in idx and row[idx["conductor"]]:
                conductor = (row[idx["conductor"]] or "").strip()
            tarifa_tercero = row[idx["tarifa_tercero"]]

            if not all([titulo, fecha_val, origen, destino, placa, tarifa_tercero]):
                raise ValueError("faltan campos obligatorios")

            fecha_servicio = fecha_val if isinstance(fecha_val, date) else date.fromisoformat(str(fecha_val))
            tarifa_tercero_num = float(tarifa_tercero)

            viaje = Viaje(
                operacion_id=operacion_id,
                tercero_id=operacion.tercero_id,
                titulo=titulo,
                fecha_servicio=fecha_servicio,
                origen=origen,
                destino=destino,
                placa=placa,
                conductor=conductor or None,
                tarifa_tercero=tarifa_tercero_num,
                created_by=user.id,
                cargado_por=user.rol.value,
            )
            viaje.tarifa_cliente, viaje.rentabilidad = calculate_tarifa_cliente(tarifa_tercero_num, operacion)
            db.add(viaje)
            cargados += 1
        except Exception as exc:
            errores.append(f"Fila {row_num}: {exc}")

    db.commit()
    return CargaMasivaResultado(total_filas=max(0, len(rows) - 1), cargados=cargados, errores=errores)
