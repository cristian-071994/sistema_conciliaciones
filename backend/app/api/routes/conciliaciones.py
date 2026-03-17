from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_current_user, is_cointra_admin
from app.db.session import get_db
from app.models.comentario import Comentario
from app.models.conciliacion import Conciliacion
from app.models.conciliacion_item import ConciliacionItem
from app.models.enums import ItemEstado, ItemTipo, UserRole
from app.models.historial_cambio import HistorialCambio
from app.models.operacion import Operacion
from app.models.usuario import Usuario
from app.models.viaje import Viaje
from app.schemas.historial import HistorialCambioOut, ResumenFinancieroOut
from app.schemas.conciliacion import (
    ClienteItemDecision,
    ComentarioCreate,
    ComentarioOut,
    ConciliacionCreate,
    ConciliacionUpdate,
    ConciliacionItemCreate,
    ConciliacionItemOut,
    ConciliacionItemPatch,
    ConciliacionItemUpdateEstado,
    ConciliacionOut,
    ConciliacionWorkflowAction,
    ConciliacionUpdateEstado,
)
from app.services.pricing import apply_rentabilidad
from app.services.audit import log_change
from app.services.notifications import create_internal_notifications, send_manual_email
from app.services.avansat import fetch_avansat_by_manifiesto
from app.services.visibility import sanitize_item_for_role
from app.schemas.viaje import AdjuntarViajesRequest, ViajeOut

router = APIRouter(prefix="/conciliaciones", tags=["conciliaciones"])


def _estado_conciliacion_viaje(viaje: Viaje):
    if not viaje.conciliacion:
        return None
    return getattr(viaje.conciliacion.estado, "value", viaje.conciliacion.estado)


def _should_mark_conciliado(estado: object) -> bool:
    estado_valor = getattr(estado, "value", estado)
    return str(estado_valor) in {"APROBADA", "CERRADA"}


def _default_viaje_item_financials(viaje: Viaje) -> tuple[float | None, float | None, float]:
    default_pct = 10.0
    tarifa_tercero = float(viaje.tarifa_tercero) if viaje.tarifa_tercero is not None else None

    if tarifa_tercero is not None:
        tarifa_cliente = tarifa_tercero / (1 - default_pct / 100)
        return tarifa_tercero, tarifa_cliente, default_pct

    tarifa_cliente = float(viaje.tarifa_cliente) if viaje.tarifa_cliente is not None else None
    if tarifa_cliente is not None:
        tarifa_tercero = tarifa_cliente * (1 - default_pct / 100)

    return tarifa_tercero, tarifa_cliente, default_pct


def _sync_viajes_conciliado_por_estado(db: Session, conciliacion_id: int, estado: object) -> None:
    estado_valor = getattr(estado, "value", estado)
    conciliado = _should_mark_conciliado(estado_valor)
    viajes = db.query(Viaje).filter(Viaje.conciliacion_id == conciliacion_id).all()
    for viaje in viajes:
        viaje.conciliado = conciliado
        viaje.estado_conciliacion = str(estado_valor)


def _repair_missing_viaje_items(db: Session, conc: Conciliacion, user_id: int) -> bool:
    """Repara inconsistencias historicas: viaje vinculado a conciliacion sin item VIAJE."""
    changed = False
    linked_viajes = db.query(Viaje).filter(Viaje.conciliacion_id == conc.id).all()
    existing_viaje_ids = {
        row[0]
        for row in db.query(ConciliacionItem.viaje_id)
        .filter(
            ConciliacionItem.conciliacion_id == conc.id,
            ConciliacionItem.tipo == ItemTipo.VIAJE,
            ConciliacionItem.viaje_id.is_not(None),
        )
        .all()
    }

    for viaje in linked_viajes:
        if viaje.id in existing_viaje_ids:
            continue

        tarifa_tercero, tarifa_cliente, rentabilidad = _default_viaje_item_financials(viaje)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            viaje_id=viaje.id,
            tipo=ItemTipo.VIAJE,
            fecha_servicio=viaje.fecha_servicio,
            origen=viaje.origen,
            destino=viaje.destino,
            placa=viaje.placa,
            conductor=viaje.conductor,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=tarifa_cliente,
            rentabilidad=rentabilidad,
            manifiesto_numero=viaje.manifiesto_numero,
            remesa=None,
            descripcion=viaje.descripcion,
            created_by=user_id,
            cargado_por=viaje.cargado_por,
        )
        db.add(item)
        log_change(
            db,
            usuario_id=user_id,
            conciliacion_id=conc.id,
            campo="reparacion_item_viaje",
            valor_nuevo=f"viaje_id={viaje.id}",
        )
        changed = True

    return changed


def _existing_item_viaje_ids(db: Session) -> set[int]:
    return {
        row[0]
        for row in db.query(ConciliacionItem.viaje_id)
        .filter(
            ConciliacionItem.tipo == ItemTipo.VIAJE,
            ConciliacionItem.viaje_id.is_not(None),
        )
        .all()
    }


def _validate_user_access_operacion(user: Usuario, operacion: Operacion) -> None:
    if user.rol == UserRole.CLIENTE and user.cliente_id != operacion.cliente_id:
        raise HTTPException(status_code=403, detail="Operacion no disponible para este cliente")
    if user.rol == UserRole.TERCERO and user.tercero_id != operacion.tercero_id:
        raise HTTPException(status_code=403, detail="Operacion no disponible para este tercero")


def _ensure_user_can_access_conciliacion(user: Usuario, conc: Conciliacion) -> None:
    if not conc.activo and not is_cointra_admin(user):
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    estado = getattr(conc.estado, "value", conc.estado)
    if user.rol == UserRole.CLIENTE and str(estado) == "BORRADOR":
        raise HTTPException(status_code=403, detail="La conciliacion aun no ha sido enviada a revision")


def _ensure_cointra_admin(user: Usuario) -> None:
    if not is_cointra_admin(user):
        raise HTTPException(status_code=403, detail="Solo COINTRA_ADMIN puede editar o inactivar conciliaciones")


def _parse_target_emails(raw_value: str | None, recipients: list[Usuario]) -> list[str]:
    provided: list[str] = []
    if raw_value:
        normalized = raw_value.replace(";", ",")
        provided = [email.strip() for email in normalized.split(",") if email and email.strip()]

    target_emails = provided or [u.email for u in recipients if u.email]
    return list(dict.fromkeys(target_emails))


def _users_matching_emails(recipients: list[Usuario], target_emails: list[str]) -> list[Usuario]:
    email_set = {email.strip().lower() for email in target_emails if email and email.strip()}
    if not email_set:
        return []
    matched: list[Usuario] = []
    seen_ids: set[int] = set()
    for user in recipients:
        if not user.email:
            continue
        if user.id in seen_ids:
            continue
        if user.email.strip().lower() in email_set:
            matched.append(user)
            seen_ids.add(user.id)
    return matched


def _sender_signature(user: Usuario) -> str:
    if user.email:
        return f"{user.nombre} <{user.email}>"
    return user.nombre


def _find_last_review_sender(db: Session, conciliacion_id: int) -> Usuario | None:
    last_sender_log = (
        db.query(HistorialCambio)
        .filter(
            HistorialCambio.conciliacion_id == conciliacion_id,
            HistorialCambio.campo == "enviar_revision",
        )
        .order_by(HistorialCambio.id.desc())
        .first()
    )
    if not last_sender_log:
        return None
    sender = db.get(Usuario, last_sender_log.usuario_id)
    if not sender or not sender.activo:
        return None
    return sender


def _resolve_recipients(db: Session, operacion: Operacion, roles: list[UserRole]) -> list[Usuario]:
    recipients: list[Usuario] = []
    for role in roles:
        query = db.query(Usuario).filter(Usuario.activo.is_(True), Usuario.rol == role)
        if role == UserRole.CLIENTE:
            query = query.filter(Usuario.cliente_id == operacion.cliente_id)
        elif role == UserRole.TERCERO:
            query = query.filter(Usuario.tercero_id == operacion.tercero_id)
        user = query.order_by(Usuario.id.asc()).first()
        if user:
            recipients.append(user)
    # Dedup por usuario
    uniq: dict[int, Usuario] = {u.id: u for u in recipients}
    return list(uniq.values())


def _display_estado(conc: Conciliacion) -> str:
    estado = str(getattr(conc.estado, "value", conc.estado))
    if estado == "APROBADA" and conc.enviada_facturacion:
        return "ENVIADA_A_FACTURAR"
    return estado


def _find_last_status_actor(db: Session, conc: Conciliacion) -> tuple[str | None, str | None]:
    estado = _display_estado(conc)
    logs = (
        db.query(HistorialCambio)
        .filter(HistorialCambio.conciliacion_id == conc.id)
        .order_by(HistorialCambio.id.desc())
        .limit(150)
        .all()
    )

    def matches(log: HistorialCambio) -> bool:
        campo = (log.campo or "").strip()
        nuevo = str(log.valor_nuevo or "").strip().upper()

        if estado == "BORRADOR":
            return campo in {"devolucion_cliente", "conciliacion_creada"} or (
                campo == "estado_conciliacion" and nuevo == "BORRADOR"
            )
        if estado == "EN_REVISION":
            return campo == "enviar_revision" or (
                campo == "estado_conciliacion" and nuevo == "EN_REVISION"
            )
        if estado == "APROBADA":
            return campo == "aprobacion_cliente" or (
                campo == "estado_conciliacion" and nuevo == "APROBADA"
            )
        if estado == "ENVIADA_A_FACTURAR":
            return campo == "envio_facturacion"
        if estado == "CERRADA":
            return campo == "cierre_conciliacion" or (
                campo == "estado_conciliacion" and nuevo == "CERRADA"
            )
        return campo == "estado_conciliacion" and nuevo == estado

    for log in logs:
        if not matches(log):
            continue
        actor = db.get(Usuario, log.usuario_id)
        if actor:
            return actor.nombre, actor.email

    creator_name = conc.creador.nombre if conc.creador else None
    creator_email = conc.creador.email if conc.creador else None
    return creator_name, creator_email


def _enrich_conciliacion(db: Session, conc: Conciliacion) -> dict:
    """Convierte una Conciliacion ORM en dict con campos de creador, cliente y tercero."""
    base = ConciliacionOut.model_validate(conc).model_dump()
    base["creador_nombre"] = conc.creador.nombre if conc.creador else None
    operacion = conc.operacion
    base["cliente_nombre"] = operacion.cliente.nombre if operacion and operacion.cliente else None
    base["tercero_nombre"] = operacion.tercero.nombre if operacion and operacion.tercero else None
    estado_actor_nombre, estado_actor_email = _find_last_status_actor(db, conc)
    base["estado_actualizado_por_nombre"] = estado_actor_nombre
    base["estado_actualizado_por_email"] = estado_actor_email
    return base


def _as_float(value: object) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def _normalize_manifiesto_for_lookup(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    # Corrige casos donde el manifiesto llega como numero decimal de Excel, p.e. "522318.0"
    if raw.endswith(".0"):
        integer_part = raw[:-2]
        if integer_part.isdigit():
            return integer_part
    return raw


def _fetch_avansat_with_fallback(manifiesto: str) -> dict:
    if not manifiesto:
        return {}
    attempts = [manifiesto, manifiesto.lstrip("0")]
    seen: set[str] = set()
    for candidate in attempts:
        candidate = (candidate or "").strip()
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        data = fetch_avansat_by_manifiesto(candidate)
        if data:
            return data
    return {}


def _prepare_facturacion_rows(items: list[ConciliacionItem]) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    missing_manifiestos: list[str] = []

    for item in items:
        viaje_ref = f"viaje #{item.viaje_id}" if item.viaje_id else f"item #{item.id}"
        manifiesto = _normalize_manifiesto_for_lookup(item.manifiesto_numero)
        if not manifiesto:
            missing_manifiestos.append(f"{viaje_ref} (sin manifiesto)")
            continue

        avansat = _fetch_avansat_with_fallback(manifiesto)
        if not avansat:
            missing_manifiestos.append(f"{viaje_ref} (manifiesto {manifiesto} sin datos en Avansat)")
            continue

        precio_cliente = _as_float(item.tarifa_cliente)
        precio_tercero = _as_float(item.tarifa_tercero)
        rentabilidad = _as_float(item.rentabilidad)
        ganancia = precio_cliente - precio_tercero

        rows.append(
            {
                "manifiesto": manifiesto,
                "fecha_emision": avansat.get("fecha_emision") or "",
                "producto": avansat.get("producto") or "",
                "placa_vehiculo": avansat.get("placa_vehiculo") or (item.placa or ""),
                "trayler": avansat.get("trayler") or "",
                "remesa": avansat.get("remesa") or (item.remesa or ""),
                "ciudad_origen": avansat.get("ciudad_origen") or (item.origen or ""),
                "ciudad_destino": avansat.get("ciudad_destino") or (item.destino or ""),
                "precio_cliente": precio_cliente,
                "precio_tercero": precio_tercero,
                "rentabilidad": rentabilidad,
                "ganancia": ganancia,
            }
        )

    return rows, sorted(set(missing_manifiestos))


def _build_facturacion_excel(conc: Conciliacion, rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturacion"

    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="FFF200")
    header_font = Font(bold=True, color="1F2937")
    total_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cop_format = '"$" #,##0'

    ws.append(
        [
            "Manifiesto",
            "Fecha Manifiesto",
            "Producto",
            "Placa",
            "Remolque",
            "Remesa",
            "Origen",
            "Destino",
            "Precio Cliente",
            "Precio Tercero",
            "Rentabilidad",
            "Ganancia Cointra",
        ]
    )

    for idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        if idx in (9, 10, 11, 12):
            cell.fill = total_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    total_precio_cliente = 0.0
    total_precio_tercero = 0.0
    total_ganancia = 0.0

    for row in rows:
        manifiesto = str(row.get("manifiesto") or "").strip()
        precio_cliente = _as_float(row.get("precio_cliente"))
        precio_tercero = _as_float(row.get("precio_tercero"))
        rentabilidad = _as_float(row.get("rentabilidad"))
        ganancia = _as_float(row.get("ganancia"))

        total_precio_cliente += precio_cliente
        total_precio_tercero += precio_tercero
        total_ganancia += ganancia

        ws.append(
            [
                manifiesto,
                row.get("fecha_emision") or "",
                row.get("producto") or "",
                row.get("placa_vehiculo") or "",
                row.get("trayler") or "",
                row.get("remesa") or "",
                row.get("ciudad_origen") or "",
                row.get("ciudad_destino") or "",
                precio_cliente,
                precio_tercero,
                rentabilidad,
                ganancia,
            ]
        )

        row_idx = ws.max_row
        for col_idx in range(1, 13):
            ws.cell(row=row_idx, column=col_idx).border = border

        for col_idx in (9, 10, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = cop_format

        ws.cell(row=row_idx, column=11).number_format = '#,##0.##" %"'

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=8, value="TOTAL")
    ws.cell(row=total_row, column=9, value=total_precio_cliente)
    ws.cell(row=total_row, column=10, value=total_precio_tercero)
    ws.cell(row=total_row, column=12, value=total_ganancia)

    for col_idx in (8, 9, 10, 12):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if col_idx in (9, 10, 12):
            cell.number_format = cop_format

    widths = {
        1: 14,
        2: 18,
        3: 24,
        4: 12,
        5: 12,
        6: 12,
        7: 16,
        8: 16,
        9: 18,
        10: 18,
        11: 14,
        12: 18,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@router.post("", response_model=ConciliacionOut)
def create_conciliacion(
    payload: ConciliacionCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    # Crear conciliaciones: COINTRA_ADMIN, COINTRA_USER (rol COINTRA)
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo usuarios Cointra pueden crear conciliaciones")

    operacion = db.get(Operacion, payload.operacion_id)
    if not operacion:
        raise HTTPException(status_code=404, detail="Operacion no encontrada")
    _validate_user_access_operacion(user, operacion)

    conc = Conciliacion(
        operacion_id=payload.operacion_id,
        nombre=payload.nombre,
        fecha_inicio=payload.fecha_inicio,
        fecha_fin=payload.fecha_fin,
        activo=True,
        enviada_facturacion=False,
        created_by=user.id,
    )
    db.add(conc)
    db.flush()

    # Cargar automaticamente todos los viajes PENDIENTES de la operacion
    viajes_pendientes = (
        db.query(Viaje)
        .filter(
            Viaje.operacion_id == payload.operacion_id,
            Viaje.conciliacion_id.is_(None),
            Viaje.fecha_servicio <= payload.fecha_fin,
        )
        .order_by(Viaje.fecha_servicio.asc(), Viaje.id.asc())
        .all()
    )

    for viaje in viajes_pendientes:
        tarifa_tercero, tarifa_cliente, rentabilidad = _default_viaje_item_financials(viaje)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            viaje_id=viaje.id,
            tipo=ItemTipo.VIAJE,
            fecha_servicio=viaje.fecha_servicio,
            origen=viaje.origen,
            destino=viaje.destino,
            placa=viaje.placa,
            conductor=viaje.conductor,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=tarifa_cliente,
            rentabilidad=rentabilidad,
            manifiesto_numero=viaje.manifiesto_numero,
            remesa=None,
            descripcion=viaje.descripcion,
            created_by=user.id,
            cargado_por=viaje.cargado_por,
        )
        estado_valor = getattr(conc.estado, "value", conc.estado)
        viaje.conciliado = _should_mark_conciliado(estado_valor)
        viaje.estado_conciliacion = str(estado_valor)
        viaje.conciliacion_id = conc.id
        db.add(item)
        log_change(
            db,
            usuario_id=user.id,
            conciliacion_id=conc.id,
            campo="viaje_adjuntado",
            valor_nuevo=f"viaje_id={viaje.id}",
        )

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="conciliacion_creada",
        valor_nuevo=f"{payload.nombre} ({payload.fecha_inicio} - {payload.fecha_fin})",
    )
    db.commit()
    db.refresh(conc)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    create_internal_notifications(
        db,
        recipients,
        titulo="Nueva conciliacion creada",
        mensaje=f"Se creo la conciliacion '{conc.nombre}' para la operacion '{operacion.nombre}'.",
        tipo="CONCILIACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc

@router.get("", response_model=list[ConciliacionOut])
def list_conciliaciones(db: Session = Depends(get_db), user: Usuario = Depends(get_current_user)):
    query = (
        db.query(Conciliacion)
        .join(Operacion, Operacion.id == Conciliacion.operacion_id)
        .options(
            selectinload(Conciliacion.operacion).selectinload(Operacion.cliente),
            selectinload(Conciliacion.operacion).selectinload(Operacion.tercero),
            selectinload(Conciliacion.creador),
        )
    )
    if user.rol == UserRole.CLIENTE and user.cliente_id:
        query = query.filter(Operacion.cliente_id == user.cliente_id, Conciliacion.estado != "BORRADOR")
    if user.rol == UserRole.TERCERO and user.tercero_id:
        query = query.filter(Operacion.tercero_id == user.tercero_id, Conciliacion.estado != "BORRADOR")
    if not is_cointra_admin(user):
        query = query.filter(Conciliacion.activo.is_(True))
    concs = query.order_by(Conciliacion.id.desc()).all()
    return [_enrich_conciliacion(db, c) for c in concs]


@router.get("/historial-cerradas", response_model=list[ConciliacionOut])
def list_closed_history(
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    cliente_id: int | None = None,
    tercero_id: int | None = None,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    query = (
        db.query(Conciliacion)
        .join(Operacion, Operacion.id == Conciliacion.operacion_id)
        .options(
            selectinload(Conciliacion.operacion).selectinload(Operacion.cliente),
            selectinload(Conciliacion.operacion).selectinload(Operacion.tercero),
            selectinload(Conciliacion.creador),
        )
        .filter(Conciliacion.estado == "CERRADA")
    )
    if user.rol == UserRole.CLIENTE and user.cliente_id:
        query = query.filter(Operacion.cliente_id == user.cliente_id)
    if user.rol == UserRole.TERCERO and user.tercero_id:
        query = query.filter(Operacion.tercero_id == user.tercero_id)
    if not is_cointra_admin(user):
        query = query.filter(Conciliacion.activo.is_(True))
    if cliente_id:
        query = query.filter(Operacion.cliente_id == cliente_id)
    if tercero_id:
        query = query.filter(Operacion.tercero_id == tercero_id)
    if fecha_inicio:
        query = query.filter(Conciliacion.fecha_inicio >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Conciliacion.fecha_fin <= fecha_fin)
    concs = query.order_by(Conciliacion.id.desc()).all()
    return [_enrich_conciliacion(db, c) for c in concs]


@router.patch("/{conciliacion_id}", response_model=ConciliacionOut)
def update_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionUpdate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No se enviaron cambios")

    if "operacion_id" in data:
        operacion = db.get(Operacion, data["operacion_id"])
        if not operacion or not operacion.activa:
            raise HTTPException(status_code=404, detail="Operacion no encontrada")

    for field, value in data.items():
        setattr(conc, field, value)

    db.commit()
    db.refresh(conc)
    return conc


@router.delete("/{conciliacion_id}")
def deactivate_conciliacion(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    conc.activo = False
    db.commit()
    return {"ok": True}


@router.post("/{conciliacion_id}/reactivar")
def reactivate_conciliacion(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    _ensure_cointra_admin(user)

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    conc.activo = True
    db.commit()
    return {"ok": True}



@router.patch("/{conciliacion_id}/estado", response_model=ConciliacionOut)
def update_estado_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionUpdateEstado,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede cambiar estado de conciliacion")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_estado = conc.estado
    conc.estado = payload.estado
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="estado_conciliacion",
        valor_anterior=old_estado,
        valor_nuevo=payload.estado,
    )
    db.commit()
    db.refresh(conc)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA, UserRole.CLIENTE, UserRole.TERCERO])
    create_internal_notifications(
        db,
        recipients,
        titulo="Cambio de estado de conciliacion",
        mensaje=f"La conciliacion '{conc.nombre}' cambio a estado {conc.estado}.",
        tipo="ESTADO",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/items", response_model=ConciliacionItemOut)
def create_item(
    payload: ConciliacionItemCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede crear items")
    conc = db.get(Conciliacion, payload.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    item = ConciliacionItem(
        conciliacion_id=payload.conciliacion_id,
        tipo=payload.tipo,
        fecha_servicio=payload.fecha_servicio,
        origen=payload.origen,
        destino=payload.destino,
        placa=payload.placa,
        conductor=payload.conductor,
        tarifa_tercero=payload.tarifa_tercero,
        tarifa_cliente=payload.tarifa_cliente,
        manifiesto_numero=payload.manifiesto_numero,
        remesa=payload.remesa,
        descripcion=payload.descripcion,
        created_by=user.id,
        cargado_por=user.rol.value,
    )

    if user.rol in [UserRole.TERCERO, UserRole.COINTRA]:
        if item.tarifa_tercero and not item.tarifa_cliente:
            apply_rentabilidad(item, operacion)

    db.add(item)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="item_creado",
        valor_nuevo=f"tipo={item.tipo}; fecha={item.fecha_servicio}",
    )
    db.commit()
    db.refresh(item)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    create_internal_notifications(
        db,
        recipients,
        titulo="Decision de cliente sobre item",
        mensaje=f"El cliente marco el item #{item.id} como {item.estado} en la conciliacion '{conc.nombre}'.",
        tipo="APROBACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return item


@router.get("/{conciliacion_id}/items", response_model=list[ConciliacionItemOut])
def list_items(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    if _repair_missing_viaje_items(db, conc, user.id):
        db.commit()

    items = (
        db.query(ConciliacionItem)
        .filter(ConciliacionItem.conciliacion_id == conciliacion_id)
        .order_by(ConciliacionItem.id.desc())
        .all()
    )

    # Enmascara campos financieros segun actor.
    return [sanitize_item_for_role(ConciliacionItemOut.model_validate(i).model_dump(), user.rol) for i in items]


@router.patch("/items/{item_id}/estado", response_model=ConciliacionItemOut)
def update_item_estado(
    item_id: int,
    payload: ConciliacionItemUpdateEstado,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede cambiar estado de items")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    conc = db.get(Conciliacion, item.conciliacion_id)
    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_estado = item.estado
    item.estado = payload.estado
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=item.id,
        campo="estado_item",
        valor_anterior=old_estado,
        valor_nuevo=payload.estado,
    )

    db.commit()
    db.refresh(item)
    return item


@router.patch("/items/{item_id}", response_model=ConciliacionItemOut)
def patch_item(
    item_id: int,
    payload: ConciliacionItemPatch,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede actualizar items")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")

    conc = db.get(Conciliacion, item.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    can_edit_borrador = conc.estado == "BORRADOR"
    can_fix_manifiesto_aprobada = conc.estado == "APROBADA" and not conc.enviada_facturacion
    if not can_edit_borrador and not can_fix_manifiesto_aprobada:
        raise HTTPException(status_code=400, detail="Solo se puede editar en BORRADOR")

    changed = payload.model_fields_set
    if can_fix_manifiesto_aprobada:
        if item.tipo != ItemTipo.VIAJE:
            raise HTTPException(status_code=400, detail="Solo viajes permiten ajustar manifiesto en estado APROBADA")
        if not changed or not changed.issubset({"manifiesto_numero"}):
            raise HTTPException(
                status_code=400,
                detail="En APROBADA solo puedes corregir el manifiesto para enviar a facturacion",
            )

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_manifiesto = item.manifiesto_numero
    old_remesa = item.remesa
    old_tarifa_tercero = item.tarifa_tercero
    old_tarifa_cliente = item.tarifa_cliente
    old_rentabilidad = item.rentabilidad

    if "manifiesto_numero" in changed:
        item.manifiesto_numero = payload.manifiesto_numero
    if "remesa" in changed:
        item.remesa = payload.remesa

    pct = float(operacion.porcentaje_rentabilidad)
    # Usar rentabilidad actual del ítem; solo como fallback la de la operación
    pct = float(item.rentabilidad) if item.rentabilidad is not None else float(operacion.porcentaje_rentabilidad)

    tarifa_fields = changed & {"tarifa_tercero", "tarifa_cliente", "rentabilidad"}
    if tarifa_fields:
        if "tarifa_tercero" in changed and "tarifa_cliente" not in changed and "rentabilidad" not in changed:
            # Modificó tarifa_tercero → recalcular tarifa_cliente; rentabilidad no cambia
            item.tarifa_tercero = payload.tarifa_tercero
            if pct < 100:
                item.tarifa_cliente = payload.tarifa_tercero / (1 - pct / 100)
        elif "tarifa_cliente" in changed and "tarifa_tercero" not in changed and "rentabilidad" not in changed:
            # Modificó tarifa_cliente → recalcular tarifa_tercero; rentabilidad no cambia
            item.tarifa_cliente = payload.tarifa_cliente
            item.tarifa_tercero = payload.tarifa_cliente * (1 - pct / 100)
        elif "rentabilidad" in changed:
            # Modificó % rentabilidad → guardar nuevo %, recalcular tarifa_tercero; tarifa_cliente no cambia
            new_pct = payload.rentabilidad if payload.rentabilidad is not None else pct
            item.rentabilidad = new_pct
            if item.tarifa_cliente is not None and new_pct < 100:
                item.tarifa_tercero = float(item.tarifa_cliente) * (1 - new_pct / 100)

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=item.id,
        campo="actualizacion_manual_item",
        valor_anterior=f"manifiesto={old_manifiesto}; remesa={old_remesa}; t3={old_tarifa_tercero}; tc={old_tarifa_cliente}; rent={old_rentabilidad}",
        valor_nuevo=f"manifiesto={item.manifiesto_numero}; remesa={item.remesa}; t3={item.tarifa_tercero}; tc={item.tarifa_cliente}; rent={item.rentabilidad}",
    )

    db.commit()
    db.refresh(item)
    return item


@router.patch("/items/{item_id}/decision-cliente", response_model=ConciliacionItemOut)
def cliente_decide_item(
    item_id: int,
    payload: ClienteItemDecision,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.CLIENTE:
        raise HTTPException(status_code=403, detail="Solo Cliente puede aprobar/rechazar items")
    if payload.estado not in [ItemEstado.APROBADO, ItemEstado.RECHAZADO]:
        raise HTTPException(status_code=400, detail="Estado permitido para Cliente: APROBADO o RECHAZADO")

    item = db.get(ConciliacionItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item no encontrado")
    conc = db.get(Conciliacion, item.conciliacion_id)
    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    old_estado = item.estado
    item.estado = payload.estado
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        item_id=item.id,
        campo="decision_cliente_item",
        valor_anterior=old_estado,
        valor_nuevo=payload.estado,
    )
    if payload.comentario:
        db.add(
            Comentario(
                conciliacion_id=conc.id,
                item_id=item.id,
                usuario_id=user.id,
                comentario=payload.comentario,
            )
        )

    db.commit()
    db.refresh(item)
    return item


@router.post("/{conciliacion_id}/enviar-revision", response_model=ConciliacionOut)
def enviar_revision(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede enviar a revision")
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    conc.estado = "EN_REVISION"
    conc.enviada_facturacion = False
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="enviar_revision",
        valor_nuevo=payload.observacion or "sin observacion",
    )

    recipients = _resolve_recipients(db, operacion, [UserRole.CLIENTE])
    target_emails = _parse_target_emails(payload.destinatario_email, recipients)
    notification_recipients = recipients

    if not target_emails:
        raise HTTPException(status_code=400, detail="No hay correo destinatario para enviar la conciliacion")

    if target_emails:
        subject = conc.nombre
        custom_message = payload.mensaje or ""
        body = (
            f"Hola,\n\n"
            f"Cointra envio la conciliacion '{conc.nombre}' para tu revision.\n"
            f"Operacion: {operacion.nombre}\n"
            f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n\n"
            f"Enviado por: {_sender_signature(user)}\n\n"
            f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
            "Ingresa al sistema para revisar y autorizar la conciliacion.\n\n"
        )
        email_result = send_manual_email(target_emails, subject=subject, body=body)
        if email_result["failed"] >= len(target_emails):
            db.rollback()
            detail = "No se pudo enviar el correo de revision"
            if email_result["errors"]:
                detail = f"{detail}: {email_result['errors'][0]}"
            raise HTTPException(status_code=502, detail=detail)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Conciliacion enviada a revision",
        mensaje=f"Cointra envio la conciliacion '{conc.nombre}' para tu revision.",
        tipo="ESTADO",
        conciliacion_id=conc.id,
    )

    db.commit()
    db.refresh(conc)
    return conc


@router.post("/{conciliacion_id}/aprobar-cliente", response_model=ConciliacionOut)
def aprobar_conciliacion_cliente(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.CLIENTE:
        raise HTTPException(status_code=403, detail="Solo Cliente puede aprobar conciliacion")
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)
    items = db.query(ConciliacionItem).filter(ConciliacionItem.conciliacion_id == conc.id).all()
    pendientes = [i for i in items if i.estado != ItemEstado.APROBADO]
    if pendientes:
        raise HTTPException(status_code=400, detail="No se puede aprobar: existen items no aprobados")

    conc.estado = "APROBADA"
    conc.enviada_facturacion = False
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="aprobacion_cliente",
        valor_nuevo=payload.observacion or "aprobada por cliente",
    )
    db.commit()
    db.refresh(conc)

    cointra_recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    tercero_recipients = _resolve_recipients(db, operacion, [UserRole.TERCERO])
    last_review_sender = _find_last_review_sender(db, conc.id)
    preferred_cointra = [last_review_sender] if last_review_sender and last_review_sender.rol == UserRole.COINTRA else []
    email_recipients = preferred_cointra or cointra_recipients
    target_emails = _parse_target_emails(payload.destinatario_email, email_recipients)
    notification_recipients = list(email_recipients)
    for tercero in tercero_recipients:
        if all(existing.id != tercero.id for existing in notification_recipients):
            notification_recipients.append(tercero)
    if target_emails:
        subject = f"Conciliacion aprobada: {conc.nombre}"
        custom_message = payload.mensaje or ""
        body = (
            f"Hola,\n\n"
            f"El cliente aprobo la conciliacion '{conc.nombre}'.\n"
            f"Operacion: {operacion.nombre}\n"
            f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n\n"
            f"Enviado por: {_sender_signature(user)}\n\n"
            f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
            "Ingresa al sistema para continuar con el flujo.\n\n"
        )
        send_manual_email(target_emails, subject=subject, body=body)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Conciliacion aprobada por cliente",
        mensaje=f"La conciliacion '{conc.nombre}' fue aprobada por el cliente y quedo autorizada para facturar.",
        tipo="APROBACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/{conciliacion_id}/devolver-cliente", response_model=ConciliacionOut)
def devolver_conciliacion_cliente(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.CLIENTE:
        raise HTTPException(status_code=403, detail="Solo Cliente puede devolver conciliacion")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    items = db.query(ConciliacionItem).filter(ConciliacionItem.conciliacion_id == conc.id).all()
    rechazados = [i for i in items if i.estado == ItemEstado.RECHAZADO]
    if not rechazados:
        raise HTTPException(
            status_code=400,
            detail="Para devolver la conciliacion debes rechazar al menos un item",
        )

    if not payload.observacion or not payload.observacion.strip():
        raise HTTPException(status_code=400, detail="Debes incluir observaciones para devolver")

    conc.estado = "BORRADOR"
    conc.enviada_facturacion = False
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="devolucion_cliente",
        valor_nuevo=payload.observacion,
    )
    db.add(
        Comentario(
            conciliacion_id=conc.id,
            usuario_id=user.id,
            comentario=payload.observacion,
        )
    )
    db.commit()
    db.refresh(conc)

    cointra_recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    last_review_sender = _find_last_review_sender(db, conc.id)
    email_recipients = [last_review_sender] if last_review_sender and last_review_sender.rol == UserRole.COINTRA else cointra_recipients
    target_emails = _parse_target_emails(payload.destinatario_email, email_recipients)
    notification_recipients = list(email_recipients)
    if target_emails:
        subject = f"Conciliacion devuelta con novedades: {conc.nombre}"
        custom_message = payload.mensaje or ""
        body = (
            f"Hola,\n\n"
            f"El cliente devolvio la conciliacion '{conc.nombre}' con novedades.\n"
            f"Operacion: {operacion.nombre}\n"
            f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n"
            f"Observacion: {payload.observacion}\n\n"
            f"Enviado por: {_sender_signature(user)}\n\n"
            f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
            "Ingresa al sistema para revisar, ajustar y reenviar.\n\n"
        )
        send_manual_email(target_emails, subject=subject, body=body)

    create_internal_notifications(
        db,
        notification_recipients,
        titulo="Conciliacion devuelta con novedades",
        mensaje=f"El cliente devolvio la conciliacion '{conc.nombre}' con observaciones para revisar.",
        tipo="DEVOLUCION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/{conciliacion_id}/enviar-facturacion", response_model=ConciliacionOut)
def enviar_facturacion_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede enviar a facturacion")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
    if conc.estado != "APROBADA":
        raise HTTPException(status_code=400, detail="Solo conciliaciones aprobadas pueden enviarse a facturacion")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    items = (
        db.query(ConciliacionItem)
        .filter(
            ConciliacionItem.conciliacion_id == conciliacion_id,
            ConciliacionItem.tipo == ItemTipo.VIAJE,
        )
        .order_by(ConciliacionItem.id.asc())
        .all()
    )
    if not items:
        raise HTTPException(status_code=400, detail="No hay viajes para generar el archivo de facturacion")

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA])
    target_emails = _parse_target_emails(payload.destinatario_email, recipients)
    if not target_emails:
        raise HTTPException(status_code=400, detail="No hay correos de destino para facturacion")

    facturacion_rows, missing_manifiestos = _prepare_facturacion_rows(items)
    if missing_manifiestos:
        count = len(missing_manifiestos)
        detail = "No se pudo generar el archivo de facturacion porque faltan datos en Avansat"
        missing_list = ", ".join(missing_manifiestos[:12])
        if len(missing_manifiestos) > 12:
            missing_list = f"{missing_list}, ..."
        raise HTTPException(
            status_code=400,
            detail=(
                f"{detail}. "
                f"Viajes pendientes ({count}): {missing_list}. "
                "Completa el manifiesto de esos viajes y vuelve a intentar."
            ),
        )

    excel_content = _build_facturacion_excel(conc, facturacion_rows)
    filename = f"conciliacion_{conc.id}_facturacion.xlsx"
    custom_message = payload.mensaje or ""
    email_body = (
        f"Hola,\n\n"
        f"Se envio la conciliacion '{conc.nombre}' para facturacion.\n"
        f"Operacion: {operacion.nombre}\n"
        f"Periodo: {conc.fecha_inicio} a {conc.fecha_fin}\n\n"
        f"Enviado por: {_sender_signature(user)}\n\n"
        f"Mensaje: {custom_message or '(sin mensaje)'}\n\n"
        "Adjunto encontraras el archivo Excel con los viajes.\n"
    )

    send_result = send_manual_email(
        target_emails,
        subject=f"Autorizacion para facturar: {conc.nombre}",
        body=email_body,
        attachments=[
            {
                "filename": filename,
                "content": excel_content,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        ],
    )
    if send_result["failed"] >= len(target_emails):
        detail = "No se pudo enviar el correo de facturacion"
        if send_result["errors"]:
            detail = f"{detail}: {send_result['errors'][0]}"
        raise HTTPException(status_code=502, detail=detail)

    conc.enviada_facturacion = True
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="envio_facturacion",
        valor_nuevo=f"destinatarios={', '.join(target_emails)}",
    )
    db.commit()
    db.refresh(conc)

    create_internal_notifications(
        db,
        recipients,
        titulo="Conciliacion enviada a facturar",
        mensaje=f"La conciliacion '{conc.nombre}' fue enviada a facturacion con archivo Excel adjunto.",
        tipo="FACTURACION",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/{conciliacion_id}/cerrar", response_model=ConciliacionOut)
def cerrar_conciliacion(
    conciliacion_id: int,
    payload: ConciliacionWorkflowAction,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede cerrar conciliacion")
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    if conc.estado != "APROBADA":
        raise HTTPException(status_code=400, detail="Solo conciliaciones aprobadas pueden cerrarse")
    operacion = db.get(Operacion, conc.operacion_id)
    conc.estado = "CERRADA"
    _sync_viajes_conciliado_por_estado(db, conc.id, conc.estado)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conc.id,
        campo="cierre_conciliacion",
        valor_nuevo=payload.observacion or "cierre formal",
    )
    db.commit()
    db.refresh(conc)

    recipients = _resolve_recipients(db, operacion, [UserRole.COINTRA, UserRole.CLIENTE, UserRole.TERCERO])
    create_internal_notifications(
        db,
        recipients,
        titulo="Conciliacion cerrada",
        mensaje=f"La conciliacion '{conc.nombre}' fue cerrada formalmente.",
        tipo="CIERRE",
        conciliacion_id=conc.id,
    )
    db.commit()
    return conc


@router.post("/comentarios", response_model=ComentarioOut)
def add_comment(
    payload: ComentarioCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol == UserRole.TERCERO:
        raise HTTPException(status_code=403, detail="Tercero no puede agregar comentarios")
    conc = db.get(Conciliacion, payload.conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)
    _ensure_user_can_access_conciliacion(user, conc)

    if payload.item_id:
        item = db.get(ConciliacionItem, payload.item_id)
        if not item or item.conciliacion_id != payload.conciliacion_id:
            raise HTTPException(status_code=400, detail="Item invalido para la conciliacion")

    comment = Comentario(
        conciliacion_id=payload.conciliacion_id,
        item_id=payload.item_id,
        usuario_id=user.id,
        comentario=payload.comentario,
    )
    db.add(comment)
    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=payload.conciliacion_id,
        item_id=payload.item_id,
        campo="comentario",
        valor_nuevo=payload.comentario,
    )
    db.commit()
    db.refresh(comment)
    return comment


@router.get("/{conciliacion_id}/comentarios", response_model=list[ComentarioOut])
def get_comments(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    return (
        db.query(Comentario)
        .filter(Comentario.conciliacion_id == conciliacion_id)
        .order_by(Comentario.id.desc())
        .all()
    )


@router.get("/{conciliacion_id}/viajes-pendientes", response_model=list[dict])
def get_pending_viajes(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo conciliaciones en BORRADOR permiten adjuntar viajes")

    already_linked_viaje_ids = _existing_item_viaje_ids(db)

    viajes = (
        db.query(Viaje)
        .filter(Viaje.operacion_id == conc.operacion_id, Viaje.conciliacion_id.is_(None))
        .order_by(Viaje.fecha_servicio.asc(), Viaje.id.asc())
        .all()
    )

    if already_linked_viaje_ids:
        viajes = [v for v in viajes if v.id not in already_linked_viaje_ids]

    payload: list[dict] = []
    for viaje in viajes:
        out = ViajeOut.model_validate(viaje).model_dump()
        out["estado_conciliacion"] = _estado_conciliacion_viaje(viaje)
        payload.append(out)

    return payload


@router.post("/{conciliacion_id}/adjuntar-viajes", response_model=list[ConciliacionItemOut])
def attach_pending_viajes(
    conciliacion_id: int,
    payload: AdjuntarViajesRequest,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede adjuntar viajes")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo conciliaciones en BORRADOR permiten adjuntar viajes")

    already_linked_viaje_ids = _existing_item_viaje_ids(db)

    viajes = (
        db.query(Viaje)
        .filter(
            Viaje.id.in_(payload.viaje_ids),
            Viaje.operacion_id == conc.operacion_id,
            Viaje.conciliacion_id.is_(None),
        )
        .all()
    )

    if already_linked_viaje_ids:
        viajes = [v for v in viajes if v.id not in already_linked_viaje_ids]

    if not viajes:
        raise HTTPException(status_code=400, detail="No hay viajes pendientes validos para adjuntar")

    created_items: list[ConciliacionItem] = []
    for viaje in viajes:
        tarifa_tercero, tarifa_cliente, rentabilidad = _default_viaje_item_financials(viaje)
        item = ConciliacionItem(
            conciliacion_id=conc.id,
            viaje_id=viaje.id,
            tipo=ItemTipo.VIAJE,
            fecha_servicio=viaje.fecha_servicio,
            origen=viaje.origen,
            destino=viaje.destino,
            placa=viaje.placa,
            conductor=viaje.conductor,
            tarifa_tercero=tarifa_tercero,
            tarifa_cliente=tarifa_cliente,
            rentabilidad=rentabilidad,
            manifiesto_numero=viaje.manifiesto_numero,
            remesa=None,
            descripcion=viaje.descripcion,
            created_by=user.id,
            cargado_por=viaje.cargado_por,
        )
        estado_valor = getattr(conc.estado, "value", conc.estado)
        viaje.conciliado = _should_mark_conciliado(estado_valor)
        viaje.estado_conciliacion = str(estado_valor)
        viaje.conciliacion_id = conc.id
        db.add(item)
        log_change(
            db,
            usuario_id=user.id,
            conciliacion_id=conc.id,
            campo="viaje_adjuntado",
            valor_nuevo=f"viaje_id={viaje.id}",
        )
        created_items.append(item)

    db.commit()
    for item in created_items:
        db.refresh(item)
    return created_items


@router.delete("/{conciliacion_id}/viajes/{viaje_id}")
def detach_viaje_from_conciliacion(
    conciliacion_id: int,
    viaje_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    if user.rol != UserRole.COINTRA:
        raise HTTPException(status_code=403, detail="Solo Cointra puede quitar viajes")

    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    if conc.estado != "BORRADOR":
        raise HTTPException(status_code=400, detail="Solo puedes quitar viajes cuando la conciliacion esta en BORRADOR")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    viaje = db.get(Viaje, viaje_id)
    if not viaje or viaje.conciliacion_id != conciliacion_id:
        raise HTTPException(status_code=404, detail="Viaje no encontrado en esta conciliacion")

    # Limpia todos los items VIAJE vinculados para ese viaje en esta conciliacion.
    (
        db.query(ConciliacionItem)
        .filter(
            ConciliacionItem.conciliacion_id == conciliacion_id,
            ConciliacionItem.tipo == ItemTipo.VIAJE,
            ConciliacionItem.viaje_id == viaje_id,
        )
        .delete(synchronize_session=False)
    )

    viaje.conciliacion_id = None
    viaje.estado_conciliacion = None
    viaje.conciliado = False

    log_change(
        db,
        usuario_id=user.id,
        conciliacion_id=conciliacion_id,
        campo="viaje_desadjuntado",
        valor_nuevo=f"viaje_id={viaje_id}",
    )

    db.commit()
    return {"ok": True}


@router.get("/{conciliacion_id}/historial", response_model=list[HistorialCambioOut])
def get_historial(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    return (
        db.query(HistorialCambio)
        .filter(HistorialCambio.conciliacion_id == conciliacion_id)
        .order_by(HistorialCambio.id.desc())
        .all()
    )


@router.get("/{conciliacion_id}/resumen-financiero", response_model=ResumenFinancieroOut)
def get_resumen_financiero(
    conciliacion_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    conc = db.get(Conciliacion, conciliacion_id)
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliacion no encontrada")

    operacion = db.get(Operacion, conc.operacion_id)
    _validate_user_access_operacion(user, operacion)

    items = db.query(ConciliacionItem).filter(ConciliacionItem.conciliacion_id == conciliacion_id).all()
    total_tercero = sum(float(i.tarifa_tercero or 0) for i in items)
    total_cliente = sum(float(i.tarifa_cliente or 0) for i in items)
    total_rentabilidad_valor = total_cliente - total_tercero
    pct_vals = [float(i.rentabilidad) for i in items if i.rentabilidad is not None]
    pct_promedio = (sum(pct_vals) / len(pct_vals)) if pct_vals else 0

    if user.rol == UserRole.COINTRA:
        return {
            "total_tarifa_tercero": total_tercero,
            "total_tarifa_cliente": total_cliente,
            "total_rentabilidad_valor": total_rentabilidad_valor,
            "total_rentabilidad_pct_promedio": pct_promedio,
        }
    if user.rol == UserRole.CLIENTE:
        return {
            "total_tarifa_tercero": None,
            "total_tarifa_cliente": total_cliente,
            "total_rentabilidad_valor": None,
            "total_rentabilidad_pct_promedio": None,
        }
    return {
        "total_tarifa_tercero": total_tercero,
        "total_tarifa_cliente": None,
        "total_rentabilidad_valor": None,
        "total_rentabilidad_pct_promedio": None,
    }
